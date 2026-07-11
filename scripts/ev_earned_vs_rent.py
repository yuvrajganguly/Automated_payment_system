r"""Per-EV: did the rider show up and earn (gross payout) enough to cover the
EV rent for the last cycle? One tab per company + a cross-company tab for riders
on 2+ companies. Engine-independent — Expected is recomputed fresh (no meter, so
no catch-up), Earned = gross PAYOUT.

    python scripts/ev_earned_vs_rent.py "postgresql://payout:payout@localhost:5432/payout"
    python scripts/ev_earned_vs_rent.py "<url>" --out "C:\path\ev_earned.xlsx"

Expected = weekly rate prorated for days the EV was held inside the cycle window,
minus maintenance days (max 0). Covered = Earned >= Expected. Missed =
max(0, Expected - Earned). Rows where the FULL weekly rate is missed are red.
"""
from __future__ import annotations

import os
import sys
from datetime import date

args = [a for a in sys.argv[1:] if not a.startswith("--")]
if not args:
    print(__doc__); sys.exit(1)
db = args[0]
out = sys.argv[sys.argv.index("--out") + 1] if "--out" in sys.argv else "ev_earned_vs_rent.xlsx"
if db.startswith("postgres://") or db.startswith("postgresql://"):
    os.environ["PAYOUT_DB_URL"] = db
else:
    os.environ["PAYOUT_DB"] = db

from payout.db import get_connection                               # noqa: E402
from payout.domain.rent import chargeable_window, maintenance_days_in_window  # noqa: E402
from payout.money import prorate                                   # noqa: E402
import openpyxl                                                    # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment           # noqa: E402

R = lambda p: round((p or 0) / 100.0, 2)  # noqa: E731
_D = lambda s: date.fromisoformat(s) if s else None  # noqa: E731


def _expected_paise(conn, ev_id, weekly, cs, ce, handover, ret):
    win = chargeable_window(_D(cs), _D(ce), _D(handover), None, _D(ret))
    if win is None:
        return 0, 0
    days = (win[1] - win[0]).days + 1
    maint = maintenance_days_in_window(conn, ev_id, win[0], win[1])
    days = max(0, days - maint)
    return prorate(int(weekly), days), days


def main():
    conn = get_connection()

    # latest completed cycle per company
    cycles = {r["company"]: (r["cycle_start"], r["cycle_end"]) for r in conn.execute(
        "SELECT cc.company, cc.cycle_start, cc.cycle_end FROM company_cycles cc "
        "JOIN (SELECT company, MAX(cycle_end) AS mx FROM company_cycles GROUP BY company) m "
        "  ON m.company=cc.company AND m.mx=cc.cycle_end")}

    # gross payout per (person, company) for each company's latest cycle
    payouts: dict = {}
    for co, (cs, ce) in cycles.items():
        for r in conn.execute(
            "SELECT person_id, COALESCE(SUM(amount),0) AS gross FROM transactions "
            "WHERE event_type='PAYOUT' AND company=? AND cycle_start=? AND cycle_end=? "
            "GROUP BY person_id", (co, cs, ce)):
            payouts[(r["person_id"], co)] = int(r["gross"] or 0)

    # person -> companies that have a cycle
    pcos: dict = {}
    for r in conn.execute("SELECT DISTINCT person_id, company FROM rider_master WHERE is_active=1"):
        if r["company"] in cycles:
            pcos.setdefault(r["person_id"], set()).add(r["company"])

    names = {r["person_id"]: r["display_name"] for r in conn.execute(
        "SELECT person_id, display_name FROM person_registry")}

    # currently-held EVs
    assigns = conn.execute(
        "SELECT ea.person_id, ea.ev_id, ea.handover_date, ea.returned_date, "
        "       m.provider, m.model_name, m.weekly_rate "
        "FROM ev_assignments ea "
        "JOIN ev_units u ON u.ev_id=ea.ev_id "
        "JOIN ev_models m ON m.model_id=u.model_id "
        "WHERE ea.returned_date IS NULL").fetchall()

    per_company: dict = {co: [] for co in cycles}
    cross: list = []
    for a in assigns:
        pid = a["person_id"]; weekly = int(a["weekly_rate"])
        cos = sorted(pcos.get(pid, set()))
        if not cos:
            continue
        holder = names.get(pid, "(unknown)")
        if len(cos) == 1:
            co = cos[0]; cs, ce = cycles[co]
            exp, days = _expected_paise(conn, a["ev_id"], weekly, cs, ce,
                                        a["handover_date"], a["returned_date"])
            earned = payouts.get((pid, co), 0)
            missed = max(0, exp - earned)
            per_company[co].append((
                a["ev_id"], a["provider"], a["model_name"], R(weekly), holder,
                "yes" if earned > 0 else "NO", days, R(exp), R(earned),
                "yes" if earned >= exp else "no", R(missed), weekly))
        else:
            ucs = min(cycles[c][0] for c in cos)
            uce = max(cycles[c][1] for c in cos)
            exp, days = _expected_paise(conn, a["ev_id"], weekly, ucs, uce,
                                        a["handover_date"], a["returned_date"])
            earned = sum(payouts.get((pid, c), 0) for c in cos)
            missed = max(0, exp - earned)
            cross.append((
                a["ev_id"], a["provider"], a["model_name"], R(weekly), holder,
                ", ".join(cos), f"{ucs}..{uce}", days, R(exp), R(earned),
                "yes" if earned >= exp else "no", R(missed), weekly))

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    RED = PatternFill("solid", fgColor="F4CCCC")

    def add(title, headers, rows, red_if_full=True, weekly_idx=None):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor="1F4E78"); c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        cov = 0
        for r in rows:
            body = r[:-1]                       # drop trailing raw weekly paise
            ws.append(list(body))
            missed = body[-1]                   # Missed ₹ is the last column
            covered = body[-2]                  # Covered? is second-last
            weekly_r = body[3]
            if (red_if_full and isinstance(missed, (int, float))
                    and missed >= weekly_r - 0.01 and missed > 0):
                for c in ws[ws.max_row]:
                    c.fill = RED
            if covered == "yes":
                cov += 1
        # money formats
        for col_letter in ("D", "H", "I", "K") if title != "Cross-company" else ("D", "I", "J", "L"):
            for cell in ws[col_letter][1:]:
                cell.number_format = "#,##0.00"; cell.alignment = Alignment(horizontal="right")
        for i, h in enumerate(headers):
            ws.column_dimensions[chr(ord("A")+i)].width = max(10, len(str(h))+2)
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{chr(ord('A')+len(headers)-1)}{len(rows)+1}"
        return cov

    co_hdr = ["EV", "Provider", "Model", "Weekly ₹", "Holder", "Present?",
              "Days", "Expected ₹", "Earned ₹", "Covered?", "Missed ₹"]
    summary = []
    for co in sorted(per_company):
        rows = sorted(per_company[co], key=lambda x: -x[10])
        cov = add(co, co_hdr, rows)
        summary.append((co, len(rows), cov, len(rows) - cov,
                        round(sum(x[10] for x in rows), 2)))
    cross_hdr = ["EV", "Provider", "Model", "Weekly ₹", "Holder", "Companies",
                 "Union window", "Days", "Expected ₹", "Combined earned ₹",
                 "Covered?", "Missed ₹"]
    if cross:
        covx = add("Cross-company", cross_hdr, sorted(cross, key=lambda x: -x[11]))
        summary.append(("Cross-company", len(cross), covx, len(cross) - covx,
                        round(sum(x[11] for x in cross), 2)))

    # Summary tab first
    sw = wb.create_sheet("Summary", 0)
    sw.append(["Company", "EVs", "Covered (earned ≥ rent)", "Not covered", "Total missed ₹"])
    for c in sw[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78"); c.font = Font(bold=True, color="FFFFFF")
    for s in summary:
        sw.append(list(s))
    sw.append(["TOTAL", sum(s[1] for s in summary), sum(s[2] for s in summary),
               sum(s[3] for s in summary), round(sum(s[4] for s in summary), 2)])
    for c in sw[sw.max_row]:
        c.font = Font(bold=True)
    for i, w in enumerate([16, 8, 22, 12, 14]):
        sw.column_dimensions[chr(ord("A")+i)].width = w

    wb.save(out)
    print(f"[ev-earned-vs-rent] {out}")
    for co, n, cov, notc, miss in summary:
        print(f"  {co:<16} EVs={n:>3}  covered={cov:>3}  not={notc:>3}  missed=₹{miss:,.2f}")


if __name__ == "__main__":
    main()
