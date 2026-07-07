r"""Backfill rider phone numbers into rider_master.mob_no.

  * Spencer's : the rider_id IS the phone number, so mob_no = rider_id.
  * Myntra    : read the 'Phone no' column from the Myntra payout files
                (Worker Code -> Phone no) and map onto riders by rider_id.

Only touches mob_no. Safe to re-run (idempotent).

    python scripts/backfill_phone_numbers.py \
        "postgresql://payout:payout@localhost:5432/payout" \
        --myntra-dir "C:\Users\Yuvraj\OneDrive\Documents\Job"

    add --dry-run to preview counts without writing.
"""
from __future__ import annotations

import glob
import os
import re
import sys

import openpyxl


def _norm_phone(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s if 7 <= len(s) <= 12 else None


def extract_myntra_phones(directory: str) -> dict[str, str]:
    """rider_id (Worker Code) -> phone, merged across every Myntra file in dir."""
    phones: dict[str, str] = {}
    for path in sorted(glob.glob(os.path.join(directory, "*.xlsx"))):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            head = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not head:
                continue
            hdr = [str(c).strip().lower() if c is not None else "" for c in head]
            rid_col = next((i for i, h in enumerate(hdr)
                            if "worker code" in h or h in ("rider_id", "rider id")), None)
            ph_col = next((i for i, h in enumerate(hdr)
                           if "phone" in h or h == "mobile" or "mob" in h), None)
            if rid_col is None or ph_col is None:
                continue
            for r in ws.iter_rows(min_row=2, values_only=True):
                if rid_col >= len(r) or ph_col >= len(r):
                    continue
                rid, ph = r[rid_col], _norm_phone(r[ph_col])
                if rid and ph and "MNOW" in str(rid):
                    phones[str(rid).strip()] = ph
        wb.close()
    return phones


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__)
        sys.exit(1)
    url = args[0]
    myntra_dir = "."
    if "--myntra-dir" in sys.argv:
        myntra_dir = sys.argv[sys.argv.index("--myntra-dir") + 1]
    dry = "--dry-run" in sys.argv

    phones = extract_myntra_phones(myntra_dir)
    print(f"[phones] extracted {len(phones)} Myntra rider->phone mappings "
          f"from {myntra_dir}")

    import psycopg
    conn = psycopg.connect(url)
    cur = conn.cursor()

    # Spencer's: rider_id == phone.
    cur.execute("SELECT COUNT(*) FROM rider_master WHERE company LIKE 'Spencer%'")
    sp_total = cur.fetchone()[0]
    if not dry:
        cur.execute(
            "UPDATE rider_master SET mob_no = rider_id "
            "WHERE company LIKE 'Spencer%' "
            "  AND (mob_no IS NULL OR mob_no <> rider_id)")

    # Myntra: map by rider_id.
    myn_updated = 0
    for rid, ph in phones.items():
        cur.execute(
            "UPDATE rider_master SET mob_no=%s "
            "WHERE company='Myntra' AND rider_id=%s AND "
            "      (mob_no IS NULL OR mob_no <> %s)",
            (ph, rid, ph))
        myn_updated += cur.rowcount

    # Coverage report for Myntra.
    cur.execute("SELECT COUNT(*) FROM rider_master WHERE company='Myntra'")
    myn_total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM rider_master WHERE company='Myntra' "
                "AND (mob_no IS NULL OR mob_no='')")
    if dry:
        conn.rollback()
        myn_missing = "?"
    else:
        conn.commit()
        myn_missing = cur.fetchone()[0]

    print(f"[Spencer's] {sp_total} riders -> mob_no = rider_id"
          + ("  (dry-run, not written)" if dry else ""))
    print(f"[Myntra]    {myn_updated} riders updated from files "
          f"({myn_total} Myntra riders total"
          + (f", {myn_missing} still without a phone)" if not dry else ")"))
    print("done." + ("  Nothing written (dry-run)." if dry else ""))


if __name__ == "__main__":
    main()
