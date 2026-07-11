r"""EV-rent report: expected vs collected (per company, latest completed cycle),
riders in arrears, inactive riders, and every manual adjustment.

Reads CURRENT state, so manual corrections are reflected. Read-only.

    python scripts/ev_rent_report.py "postgresql://payout:payout@localhost:5432/payout"
    python scripts/ev_rent_report.py "<url>" --out "C:\path\report.xlsx"
"""
from __future__ import annotations

import os
import sys
from datetime import date

args = [a for a in sys.argv[1:] if not a.startswith("--")]
if not args:
    print(__doc__)
    sys.exit(1)
db = args[0]
out = "ev_rent_report.xlsx"
if "--out" in sys.argv:
    out = sys.argv[sys.argv.index("--out") + 1]
if db.startswith("postgres://") or db.startswith("postgresql://"):
    os.environ["PAYOUT_DB_URL"] = db
else:
    os.environ["PAYOUT_DB"] = db

from payout.db import get_connection  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402

R = lambda p: round((p or 0) / 100.0, 2)  # paise -> rupees  # noqa: E731
HEAD = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(bold=True, color="FFFFFF")
MONEY = "#,##0.00"


def _sheet(wb, title, headers, rows, money_cols=(), total_cols=()):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEAD; c.font = HFONT; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r))
    if rows and total_cols:
        tot = [""] * len(headers)
        tot[0] = "TOTAL"
        for i in total_cols:
            tot[i] = sum((r[i] or 0) for r in rows)
        ws.append(tot)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
    for i in money_cols:
        col = chr(ord("A") + i)
        for cell in ws[col][1:]:
            cell.number_format = MONEY
            cell.alignment = Alignment(horizontal="right")
    for i, h in enumerate(headers):
        ws.column_dimensions[chr(ord("A") + i)].width = max(12, len(str(h)) + 3)
    ws.freeze_panes = "A2"
    if len(rows) > 1:
        ws.auto_filter.ref = f"A1:{chr(ord('A')+len(headers)-1)}{len(rows)+1}"
    return ws


def main():
    conn = get_connection()

    # 1) Latest completed cycle per company -> expected vs collected.
    cycles = conn.execute(
        "SELECT cc.company, cc.cycle_start, cc.cycle_end "
        "FROM company_cycles cc "
        "JOIN (SELECT company, MAX(cycle_end) AS mx FROM company_cycles GROUP BY company) m "
        "  ON m.company = cc.company AND m.mx = cc.cycle_end "
        "ORDER BY cc.company"
    ).fetchall()
    evc_rows = []
    for c in cycles:
        agg = conn.execute(
            "SELECT "
            " SUM(CASE WHEN event_type='RENT' THEN -amount ELSE 0 END) AS charged, "
            " SUM(CASE WHEN event_type='RENT_COLLECTED' THEN amount ELSE 0 END) AS collected, "
            " SUM(CASE WHEN event_type='RENT_MISSED' THEN -amount ELSE 0 END) AS missed, "
            " SUM(CASE WHEN event_type IN ('RENT_RECOVERED','XC_RENT_RECOVERED') "
            "          THEN amount ELSE 0 END) AS recovered "
            "FROM transactions WHERE company=? AND cycle_start=? AND cycle_end=?",
            (c["company"], c["cycle_start"], c["cycle_end"])).fetchone()
        charged, collected = R(agg["charged"]), R(agg["collected"])
        missed, recovered = R(agg["missed"]), R(agg["recovered"])
        expected = round(charged + missed, 2)              # billed + missed = due this cycle
        shortfall = round(expected - collected, 2)         # rolled-to-dues + still-missed
        rate = round(100 * collected / expected, 1) if expected else 0.0
        evc_rows.append((c["company"], f"{c['cycle_start']}..{c['cycle_end']}",
                         expected, collected, shortfall, rate, recovered))

    # 2) Arrears (current, netted against credit balance = Total Dues).
    arr_rows = [
        (r["display_name"], r["deduction_company"] or "-", R(r["outstanding"]),
         R(r["bal"]), R(r["total_dues"]))
        for r in conn.execute(
            "SELECT pr.display_name, pr.deduction_company, ea.outstanding, "
            "       COALESCE(b.current_balance,0) AS bal, "
            "       ea.outstanding - COALESCE(b.current_balance,0) AS total_dues "
            "FROM ev_arrears ea "
            "JOIN person_registry pr ON pr.person_id = ea.person_id "
            "LEFT JOIN balances b ON b.person_id = ea.person_id "
            "WHERE ea.outstanding - COALESCE(b.current_balance,0) > 0 "
            "ORDER BY total_dues DESC")]

    # 3) Inactive: still holding an EV AND (owe arrears OR flagged inactive).
    inact_rows = [
        (r["display_name"], r["deduction_company"] or "-", r["ev_id"],
         r["rent_charged_through"], R(r["outstanding"]), r["status"], r["last_seen"])
        for r in conn.execute(
            "SELECT pr.display_name, pr.deduction_company, ea2.ev_id, "
            "       ea2.rent_charged_through, COALESCE(ar.outstanding,0) AS outstanding, "
            "       COALESCE(st.status,'?') AS status, st.last_seen "
            "FROM person_registry pr "
            "JOIN ev_assignments ea2 ON ea2.person_id=pr.person_id AND ea2.returned_date IS NULL "
            "LEFT JOIN ev_arrears ar ON ar.person_id=pr.person_id "
            "LEFT JOIN status_tracking st ON st.person_id=pr.person_id "
            "WHERE COALESCE(ar.outstanding,0) > 0 OR COALESCE(st.status,'')='inactive' "
            "ORDER BY outstanding DESC")]

    # 4) Manual adjustments (transparency for the manual corrections).
    adj_rows = [
        (r["created_at"], r["display_name"], r["company"] or "-", R(r["amount"]),
         r["created_by"], r["remarks"])
        for r in conn.execute(
            "SELECT t.created_at, pr.display_name, t.company, t.amount, "
            "       t.created_by, t.remarks "
            "FROM transactions t JOIN person_registry pr ON pr.person_id=t.person_id "
            "WHERE t.event_type='ADJUSTMENT' ORDER BY t.id DESC")]

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "Expected vs Collected",
           ["Company", "Latest cycle", "Expected (rent due)", "Collected (this cycle)",
            "Shortfall", "Collected %", "Old arrears recovered"],
           evc_rows, money_cols=(2, 3, 4, 6), total_cols=(2, 3, 4, 6))
    _sheet(wb, "Arrears",
           ["Rider", "Company", "EV arrears", "Credit balance", "Total dues"],
           arr_rows, money_cols=(2, 3, 4), total_cols=(2, 4))
    _sheet(wb, "Inactive (holding EV, owing)",
           ["Rider", "Company", "EV", "Rent through", "Arrears", "Status", "Last seen"],
           inact_rows, money_cols=(4,), total_cols=(4,))
    _sheet(wb, "Manual adjustments",
           ["When", "Rider", "Company", "Amount", "By", "Reason"],
           adj_rows, money_cols=(3,), total_cols=(3,))
    wb.save(out)
    print(f"[report] {out}")
    print(f"  companies (latest cycle): {len(evc_rows)}")
    print(f"  riders in arrears       : {len(arr_rows)}")
    print(f"  inactive (holding EV)   : {len(inact_rows)}")
    print(f"  manual adjustments      : {len(adj_rows)}")


if __name__ == "__main__":
    main()
