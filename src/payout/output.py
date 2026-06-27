"""Excel output builder for a CycleResult."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from payout.db import get_connection
from payout.money import to_rupees

HDR_FG = "1F4E79"; HDR_TX = "FFFFFF"
HOLD_FILL = "FFD966"; DUES_FILL = "FCE4D6"; FLAG_FILL = "FF7F7F"
PAY_FILL = "E2EFDA"; TOTAL_FILL = "D9D9D9"
EVENT_COLORS = {
    "PAYOUT": "E2EFDA", "RENT": "FCE4D6", "RENT_MISSED": "FF7F7F",
    "RENT_RECOVERED": "DAE8FC", "DUES_CARRY": "FFF2CC", "ADJUSTMENT": "E1D5E7",
    "DEDUCTION_SWITCH": "D5E8D4", "EV_SWAP": "D9E1F2", "OPENING": "F5F5F5",
}
FONT_NAME = "Arial"
NUM = "#,##0.00"   # plain thousands + 2dp, no currency symbol


def _hdr_font(): return Font(name=FONT_NAME, bold=True, color=HDR_TX, size=10)
def _body_font(bold=False): return Font(name=FONT_NAME, bold=bold, size=10)
def _fill(c): return PatternFill("solid", fgColor=c)
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _hdr_font(); c.fill = _fill(HDR_FG)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = _border()
    if row == 1: ws.freeze_panes = "A2"


def _style_row(ws, row_idx, n_cols, fill=None, bold=False):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = _body_font(bold=bold); c.border = _border()
        if fill: c.fill = _fill(fill)


def _set_num(ws, row_idx, cols):
    """Format the given (money) columns as INR and convert stored paise to
    rupees. SUM-formula cells (strings) are left alone; they sum the already
    converted rupee cells above them."""
    for col in cols:
        c = ws.cell(row=row_idx, column=col)
        if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
            c.value = to_rupees(c.value)
        c.number_format = NUM


def _auto_width(ws, min_w=10, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(min_w, min(max_w, max(
            (len(str(c.value)) if c.value is not None else 0) for c in col) + 4))
        ws.column_dimensions[letter].width = width


# ---- PAY ----
PAY_HEADERS = ["Person ID","Rider ID","Name","Hub","Vehicle","EV ID","Orders",
               "Rent Charged","Gross Payout","Previous Dues","Dues Cleared",
               "Total Deductions","Net Release","Carry Forward","COD Hold",
               "Remarks","Account No","IFSC","Manual Adjustment","Notes"]
# Numeric (right-align + thousands separator) columns. Orders=7 stays integer-ish
# but we leave it in the numeric set so it formats consistently.
PAY_NUM = [8, 9, 10, 11, 12, 13, 14, 15, 19]   # money only (Orders=7 excluded)
PAY_SUM = [7, 8, 9, 10, 11, 12, 13, 14, 15]


def _pay_sheet(wb, rows):
    ws = wb.create_sheet("PAY")
    _write_header(ws, PAY_HEADERS)
    for i, r in enumerate(rows, 2):
        # Previous Dues = prior general dues + prior EV-rent arrears
        # (arrears outstanding coming into the cycle = new_arrears + what
        # was recovered this cycle).
        prev_arrears = (r.new_arrears or 0.0) + (r.arrears_recovered or 0.0)
        prev_dues = round(max(0.0, -r.prev_balance) + prev_arrears, 2)
        carry = round(max(0.0, -r.new_balance), 2)
        # Dues Cleared = whatever portion of the payout retired carried-forward
        # dues this cycle. The engine surfaces this in dues_cleared so the row
        # breakdown is rent + arrears_recovered + dues_cleared = total deduction.
        dues_cleared = round(getattr(r, "dues_cleared", 0.0) or 0.0, 2)
        deductions = round(r.payout - r.released, 2)
        orders_val = getattr(r, "orders", None)
        fill = HOLD_FILL if r.is_hold else PAY_FILL
        vals = [r.person_id, r.rider_id, r.name, r.hub or "", r.vehicle or "",
                r.ev_id or "", orders_val if orders_val is not None else "",
                r.rent, r.payout, prev_dues, dues_cleared, deductions, r.released,
                carry, r.cod_hold, r.remarks, r.account_no or "", r.ifsc or "",
                "", ""]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            # Left-align Name (3), Remarks (16), Account/IFSC (17/18), Notes (20).
            cell.alignment = Alignment(
                horizontal="left" if col in (3, 16, 17, 18, 20) else "center",
                vertical="center",
            )
        _style_row(ws, i, len(PAY_HEADERS), fill=fill)
        _set_num(ws, i, PAY_NUM)
    if rows:
        tr = len(rows) + 2
        ws.cell(row=tr, column=2, value="TOTAL")
        for col in PAY_SUM:
            letter = get_column_letter(col)
            ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr-1})")
        _style_row(ws, tr, len(PAY_HEADERS), fill=TOTAL_FILL, bold=True)
        _set_num(ws, tr, PAY_NUM)
    _auto_width(ws)


# ---- DUES ----
DUES_HEADERS = ["Person ID","Rider ID","Name","Hub","Vehicle","EV ID","Model",
                "Orders","Gross Payout","Rent Charged","Arrears Recovered",
                "Dues Cleared","Previous Dues","Carry Forward","Account No","IFSC"]
DUES_NUM = [9, 10, 11, 12, 13, 14]             # money only (Orders=8 excluded)
DUES_SUM = [8, 9, 10, 11, 12, 13, 14]


def _dues_sheet(wb, rows):
    ws = wb.create_sheet("DUES")
    _write_header(ws, DUES_HEADERS)
    for i, r in enumerate(rows, 2):
        # Previous Dues = prior general dues + prior EV-rent arrears
        # (arrears outstanding coming into the cycle = new_arrears + what
        # was recovered this cycle).
        prev_arrears = (r.new_arrears or 0.0) + (r.arrears_recovered or 0.0)
        prev_dues = round(max(0.0, -r.prev_balance) + prev_arrears, 2)
        carry = round(max(0.0, -r.new_balance), 2)
        dues_cleared = round(getattr(r, "dues_cleared", 0.0) or 0.0, 2)
        orders_val = getattr(r, "orders", None)
        vals = [r.person_id, r.rider_id, r.name, r.hub or "", r.vehicle or "",
                r.ev_id or "", r.model or "",
                orders_val if orders_val is not None else "",
                r.payout, r.rent, r.arrears_recovered, dues_cleared, prev_dues,
                carry, r.account_no or "", r.ifsc or ""]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v).alignment = Alignment(
                horizontal="left" if col in (3, 15, 16) else "center", vertical="center")
        _style_row(ws, i, len(DUES_HEADERS), fill=DUES_FILL)
        _set_num(ws, i, DUES_NUM)
    if rows:
        tr = len(rows) + 2
        ws.cell(row=tr, column=2, value="TOTAL")
        for col in DUES_SUM:
            letter = get_column_letter(col)
            ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr-1})")
        _style_row(ws, tr, len(DUES_HEADERS), fill=TOTAL_FILL, bold=True)
        _set_num(ws, tr, DUES_NUM)
    _auto_width(ws)


# ---- ARREARS ----
ARR_HEADERS = ["Person ID","Name","EV ID","Model",
               "Total Missed","Total Recovered","Recovered This Cycle","Outstanding"]
ARR_NUM = [5, 6, 7, 8]


def _arrears_sheet(wb, conn, cs, ce):
    ws = wb.create_sheet("ARREARS")
    _write_header(ws, ARR_HEADERS)
    rows = conn.execute("""
        SELECT pr.person_id, pr.display_name, a.ev_id, m.model_name,
               COALESCE(ea.total_missed,0)    AS tm,
               COALESCE(ea.total_recovered,0) AS tr,
               COALESCE(ea.outstanding,0)     AS out_,
               COALESCE((SELECT SUM(amount) FROM transactions t
                         WHERE t.person_id=pr.person_id AND t.event_type='RENT_RECOVERED'
                         AND t.cycle_start=? AND t.cycle_end=?), 0) AS rec_cycle
        FROM person_registry pr
        LEFT JOIN ev_arrears   ea ON ea.person_id = pr.person_id
        LEFT JOIN ev_assignments a ON a.person_id = pr.person_id AND a.returned_date IS NULL
        LEFT JOIN ev_units      u ON u.ev_id = a.ev_id
        LEFT JOIN ev_models     m ON m.model_id = u.model_id
        WHERE COALESCE(ea.outstanding,0) > 0
           OR COALESCE(ea.total_missed,0) > 0
           OR COALESCE(ea.total_recovered,0) > 0
        ORDER BY ea.outstanding DESC
    """, (cs.isoformat(), ce.isoformat())).fetchall()
    for i, r in enumerate(rows, 2):
        vals = [r["person_id"], r["display_name"], r["ev_id"] or "", r["model_name"] or "",
                r["tm"], r["tr"], r["rec_cycle"], r["out_"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v).alignment = Alignment(
                horizontal="left" if col == 2 else "center", vertical="center")
        fill = FLAG_FILL if r["out_"] > 0 else None
        _style_row(ws, i, len(ARR_HEADERS), fill=fill)
        _set_num(ws, i, ARR_NUM)
    _auto_width(ws)


# ---- HOLD ----
HOLD_HEADERS = ["Rider ID", "Name", "COD Total"]
LINE_HEADERS = ["Rider ID", "Order Number", "Amount", "Payment Mode", "Status", "Source"]


def _hold_sheet(wb, conn, company, cs, ce):
    ws = wb.create_sheet("HOLD")
    _write_header(ws, HOLD_HEADERS)
    rows = conn.execute("""
        SELECT ch.rider_id, MAX(rm.name) AS name, SUM(ch.amount) AS total
        FROM cod_holds ch
        LEFT JOIN rider_master rm ON rm.rider_id = ch.rider_id AND rm.company = ch.company
        WHERE ch.company=? AND ch.cycle_start=? AND ch.cycle_end=?
        GROUP BY ch.rider_id ORDER BY total DESC
    """, (company, cs.isoformat(), ce.isoformat())).fetchall()
    for i, r in enumerate(rows, 2):
        vals = [r["rider_id"], r["name"] or "", r["total"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v).alignment = Alignment(
                horizontal="left" if col == 2 else "center", vertical="center")
        _style_row(ws, i, len(HOLD_HEADERS), fill=HOLD_FILL)
        _hc = ws.cell(row=i, column=3)
        if isinstance(_hc.value,(int,float)) and not isinstance(_hc.value,bool): _hc.value = to_rupees(_hc.value)
        _hc.number_format = NUM
    start = len(rows) + 4
    if start > 4:
        ws.cell(row=start - 1, column=1, value="Line Items").font = _body_font(bold=True)
    _write_header(ws, LINE_HEADERS, row=start)
    line_rows = conn.execute("""
        SELECT rider_id, order_number, amount, payment_mode, txn_status, source
        FROM cod_holds WHERE company=? AND cycle_start=? AND cycle_end=?
        ORDER BY rider_id, amount DESC
    """, (company, cs.isoformat(), ce.isoformat())).fetchall()
    for i, r in enumerate(line_rows, start + 1):
        vals = [r["rider_id"], r["order_number"] or "", r["amount"],
                r["payment_mode"] or "", r["txn_status"] or "", r["source"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v).alignment = Alignment(horizontal="center")
        _style_row(ws, i, len(LINE_HEADERS))
        _hc = ws.cell(row=i, column=3)
        if isinstance(_hc.value,(int,float)) and not isinstance(_hc.value,bool): _hc.value = to_rupees(_hc.value)
        _hc.number_format = NUM
    _auto_width(ws)


# ---- INACTIVE ----
INA_HEADERS = ["Person ID","Name","Hub","Rider IDs at Company","Vehicle","EV ID",
               "Model","Current Balance","Arrears Outstanding","Reason"]
INA_NUM = [8, 9]


def _inactive_sheet(wb, rows):
    ws = wb.create_sheet("INACTIVE")
    _write_header(ws, INA_HEADERS)
    for i, r in enumerate(rows, 2):
        # Vehicle defaults to BIKE so any older row without an explicit value
        # still shows something useful.
        veh = (getattr(r, "vehicle", None) or "BIKE")
        hub = (getattr(r, "hub", None) or "")
        vals = [r.person_id, r.name, hub, ", ".join(r.rider_ids), veh,
                r.ev_id or "", r.model or "",
                r.current_balance, r.arrears_outstanding, r.reason]
        for col, v in enumerate(vals, 1):
            # Left-align Name (2), Hub (3), Rider IDs list (4), Reason (10).
            ws.cell(row=i, column=col, value=v).alignment = Alignment(
                horizontal="left" if col in (2, 3, 4, 10) else "center",
                vertical="center")
        flagged = r.current_balance < 0 or r.arrears_outstanding > 0 or "Missed" in (r.reason or "")
        _style_row(ws, i, len(INA_HEADERS), fill=FLAG_FILL if flagged else None)
        _set_num(ws, i, INA_NUM)
    _auto_width(ws)


# ---- AUDIT ----
AUDIT_HEADERS = ["Person ID","Rider ID","Company","Cycle Start","Cycle End",
                 "Event Type","Amount","Balance After","Days","Remarks","Created By"]
AUDIT_NUM = [7, 8]


def _audit_sheet(wb, conn, cs, ce):
    ws = wb.create_sheet("AUDIT")
    _write_header(ws, AUDIT_HEADERS)
    rows = conn.execute("""
        SELECT person_id, rider_id, company, cycle_start, cycle_end,
               event_type, amount, balance_after, days, remarks, created_by
        FROM transactions WHERE cycle_start=? AND cycle_end=? ORDER BY id
    """, (cs.isoformat(), ce.isoformat())).fetchall()
    for i, r in enumerate(rows, 2):
        vals = [r["person_id"], r["rider_id"], r["company"], r["cycle_start"], r["cycle_end"],
                r["event_type"], r["amount"], r["balance_after"], r["days"],
                r["remarks"], r["created_by"]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=v).alignment = Alignment(horizontal="center")
        _style_row(ws, i, len(AUDIT_HEADERS), fill=EVENT_COLORS.get(r["event_type"]))
        _set_num(ws, i, AUDIT_NUM)
    _auto_width(ws)


def build_output(cycle_result) -> BytesIO:
    """Build the styled .xlsx for a committed CycleResult."""
    wb = Workbook(); wb.remove(wb.active)
    conn = get_connection()
    try:
        _pay_sheet(wb, cycle_result.pay_rows)
        _dues_sheet(wb, cycle_result.dues_rows)
        _arrears_sheet(wb, conn, cycle_result.cycle_start, cycle_result.cycle_end)
        _hold_sheet(wb, conn, cycle_result.company, cycle_result.cycle_start, cycle_result.cycle_end)
        _inactive_sheet(wb, cycle_result.inactive_rows)
        _audit_sheet(wb, conn, cycle_result.cycle_start, cycle_result.cycle_end)
    finally:
        conn.close()
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf


def build_output_filename(company: str, cycle_start: date, cycle_end: date) -> str:
    return (f"payout_{company.lower()}_{cycle_start.strftime('%d%b%Y')}_"
            f"{cycle_end.strftime('%d%b%Y')}.xlsx").replace(" ", "_")
