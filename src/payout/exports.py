"""Generic table → styled .xlsx StreamingResponse builder.

Used by the per-page export endpoints (riders, EVs, arrears, COD, etc.) so
every download has the same header bar, freeze pane, numeric formatting, and
auto-width treatment as the cycle output workbook. One helper avoids each
endpoint reinventing openpyxl boilerplate.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from datetime import datetime
from io import BytesIO

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from payout.money import to_rupees

# Same palette as payout/output.py.
HDR_FG = "1F4E79"
HDR_TX = "FFFFFF"
TOTAL_FILL = "D9D9D9"
FONT_NAME = "Arial"
NUM = "#,##0.00"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _hdr_font():
    return Font(name=FONT_NAME, bold=True, color=HDR_TX, size=10)


def _body_font(bold: bool = False):
    return Font(name=FONT_NAME, bold=bold, size=10)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws, min_w: int = 10, max_w: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(
            min_w,
            min(max_w, max((len(str(c.value)) if c.value is not None else 0) for c in col) + 4),
        )
        ws.column_dimensions[letter].width = width


def _rupeeize_rows(rows: list, money_cols: Sequence[int]) -> list:
    """Convert the 1-based ``money_cols`` of every row from paise to rupees."""
    if not money_cols:
        return rows
    mset = set(money_cols)
    return [
        [
            (
                to_rupees(v)
                if (i + 1) in mset and isinstance(v, (int, float)) and not isinstance(v, bool)
                else v
            )
            for i, v in enumerate(r)
        ]
        for r in rows
    ]


def build_xlsx(
    *,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    numeric_cols: Sequence[int] = (),
    totals_cols: Sequence[int] = (),
    left_align_cols: Sequence[int] = (),
    money_cols: Sequence[int] = (),
) -> BytesIO:
    """Build a one-sheet styled .xlsx in memory.

    Args:
        sheet_name:        Excel sheet name (kept short — Excel max is 31 chars).
        headers:           Column headers, in display order.
        rows:              Iterable of row tuples (same length as headers).
        numeric_cols:      1-based column indexes to format as ``#,##0.00`` and
                           right-align.
        totals_cols:       1-based columns to add a SUM() footer for.
        left_align_cols:   1-based columns that should be left-aligned (e.g.
                           free-form text like names / remarks).
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name[:31])
    numeric_set = set(numeric_cols)
    left_set = set(left_align_cols)

    # Header row.
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _hdr_font()
        c.fill = _fill(HDR_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws.freeze_panes = "A2"

    # Body. Money columns are stored as integer paise internally -> rupees here.
    rows = _rupeeize_rows(list(rows), money_cols)
    for i, row in enumerate(rows, 2):
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = _body_font()
            c.border = _border()
            if col in left_set:
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif col in numeric_set:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # TOTAL footer.
    if rows and totals_cols:
        tr = len(rows) + 2
        ws.cell(row=tr, column=1, value="TOTAL")
        for col in totals_cols:
            letter = get_column_letter(col)
            ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr - 1})")
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=tr, column=col)
            c.font = _body_font(bold=True)
            c.border = _border()
            c.fill = _fill(TOTAL_FILL)
            if col in numeric_set:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def add_styled_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    numeric_cols: Sequence[int] = (),
    totals_cols: Sequence[int] = (),
    left_align_cols: Sequence[int] = (),
    money_cols: Sequence[int] = (),
) -> None:
    """Append one styled sheet to an existing Workbook.

    Same look as ``build_xlsx`` but works on a multi-sheet report. Skips the
    sheet entirely (creating an empty one with just headers) when ``rows`` is
    empty — keeps the report's sheet structure stable so consumers don't have
    to handle missing tabs.
    """
    ws = wb.create_sheet(sheet_name[:31])
    numeric_set = set(numeric_cols)
    left_set = set(left_align_cols)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _hdr_font()
        c.fill = _fill(HDR_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws.freeze_panes = "A2"

    # Money columns are stored as integer paise internally -> rupees here.
    # (This parameter was declared and never used: 8 of the dashboard export's
    # 9 sheets showed paise as if they were rupees.)
    rows = _rupeeize_rows(list(rows), money_cols)
    for i, row in enumerate(rows, 2):
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = _body_font()
            c.border = _border()
            if col in left_set:
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif col in numeric_set:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    if rows and totals_cols:
        tr = len(rows) + 2
        ws.cell(row=tr, column=1, value="TOTAL")
        for col in totals_cols:
            letter = get_column_letter(col)
            ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr - 1})")
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=tr, column=col)
            c.font = _body_font(bold=True)
            c.border = _border()
            c.fill = _fill(TOTAL_FILL)
            if col in numeric_set:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    _auto_width(ws)


def workbook_response(wb: Workbook, filename_stem: str) -> Response:
    """Serialize a (multi-sheet) Workbook to a FastAPI Response download."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"{filename_stem}_{stamp}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


def xlsx_response(
    *,
    filename_stem: str,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    numeric_cols: Sequence[int] = (),
    totals_cols: Sequence[int] = (),
    left_align_cols: Sequence[int] = (),
    money_cols: Sequence[int] = (),
) -> Response:
    """Wrap ``build_xlsx`` in a FastAPI Response with proper headers.

    Filename is suffixed with the current timestamp so repeated downloads
    don't clobber each other in the browser's download folder.
    """
    buf = build_xlsx(
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
        numeric_cols=numeric_cols,
        totals_cols=totals_cols,
        left_align_cols=left_align_cols,
        money_cols=money_cols,
    )
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    name = f"{filename_stem}_{stamp}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )
