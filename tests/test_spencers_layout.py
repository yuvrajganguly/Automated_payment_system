"""Spencer's 2026-08 payout layout, hub sync from the file, and the HOLD sheet.

The client's export changed shape: the payout sheet is keyed on ``rider_phone``
(the rider id has always been the phone number), pays ``Total Payable`` and
names the store in ``store_names``; the COD sheet is called "COD HOLD" and
carries ``HUB CODE`` / ``WORKER NAME`` / ``Transaction Status``.

Three behaviours are pinned here:
1. the parser reads both the old and the new layout with the same config;
2. every rider in the payout has their roster hub rewritten from the file, and
   the PAY/DUES rows show the file's hub, not the roster's;
3. the HOLD sheet lists COD riders who are NOT in the payout in their own block,
   with the hub the COD sheet gave, and the Line Items carry name + hub.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook, load_workbook

from payout.domain.engine import process_cycle
from payout.output import NOT_IN_PAYOUT_TITLE, build_output
from payout.parsers import parse_file
from tests.conftest import make_person, make_rider

_NEW_PAYOUT_HDR = [
    "rider_phone",
    "rider_name",
    "rider_type",
    "cluster_names",
    "store_ids",
    "store_names",
    "total_orders_delivered",
    "Base earning",
    "MG",
    "Total Payable",
]
_NEW_COD_HDR = [
    "ORDER NUMBER",
    "HUB CODE",
    "WORKER CODE",
    "WORKER NAME",
    "AMOUNT",
    "Transaction Status",
    "Transaction Type",
]


def _new_layout(payout_rows, cod_rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payout"
    ws.append(_NEW_PAYOUT_HDR)
    for r in payout_rows:
        ws.append(r)
    cod = wb.create_sheet("COD HOLD")
    cod.append(_NEW_COD_HDR)
    for r in cod_rows:
        cod.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _old_layout(payout_rows, cod_rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "WEEK1"
    ws.append(["Rider id", "Rider Name", "Store", "Delivered Orders", "Total Payable Amount"])
    for r in payout_rows:
        ws.append(r)
    cod = wb.create_sheet("COD")
    cod.append(["ORDER NUMBER", "WORKER CODE", "AMOUNT"])
    for r in cod_rows:
        cod.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 1. parser ────────────────────────────────────────────────────────────────


def test_parser_reads_new_spencers_layout(db):
    data = _new_layout(
        [
            [
                6295515978,
                "Pradip Ray",
                "FULL_TIME",
                "Tamal",
                "h069",
                "NTS",
                300,
                10500,
                18650,
                18825,
            ],
            [
                8355007139,
                "Rajbir",
                "FULL_TIME",
                "Gorakhpur",
                "s055",
                "Padri bazaar",
                282,
                14100,
                0,
                14380,
            ],
        ],
        [
            [3008273151, "H049", 7439823275, "Anish Biswas", 609, "Pending", "COD"],
            [3008272746, "H069", 6295515978, "Pradip Ray", 349, "Pending", "COD"],
            [3008272747, "H069", 6295515978, "Pradip Ray", 100, "Settled", "COD"],
        ],
    )
    res = parse_file("Spencer's", data)
    assert res.sheet == "Payout"
    assert res.matched_columns["rider_id"] == "rider_phone"
    assert res.matched_columns["payout"] == "Total Payable"
    assert res.matched_columns["orders"] == "total_orders_delivered"
    assert res.matched_columns["hub"] == "store_names"
    assert res.matched_columns["hold_sheet"] == "COD HOLD"
    by_id = {r.rider_id: r for r in res.records}
    assert set(by_id) == {"6295515978", "8355007139"}
    assert by_id["6295515978"].payout == 18825 * 100
    assert by_id["6295515978"].orders == 300
    assert by_id["6295515978"].name == "Pradip Ray"
    assert by_id["6295515978"].hub == "NTS"
    # COD lines: numeric worker codes normalised like rider ids, hub + name kept,
    # and the status column picked up without any config for it.
    assert [(c.worker_code, c.hub, c.name, c.txn_status) for c in res.cod_lines] == [
        ("7439823275", "H049", "Anish Biswas", "Pending"),
        ("6295515978", "H069", "Pradip Ray", "Pending"),
        ("6295515978", "H069", "Pradip Ray", "Settled"),
    ]


def test_phone_style_rider_ids_are_normalised():
    from payout.parsers.generic import _normalise_rider_id as n

    # Hand-typed COD rows carry the phone in every spelling; all must match the
    # payout sheet's plain 10-digit number.
    assert n("'+916291274901") == "6291274901"
    assert n("+91 62912 74901") == "6291274901"
    assert n("06291274901") == "6291274901"
    assert n(6291274901) == "6291274901"
    assert n(6291274901.0) == "6291274901"
    assert n("6291274901.0") == "6291274901"
    # Non-phone ids are untouched.
    assert n("BD-12") == "BD-12"
    assert n(" J1 ") == "J1"
    assert n("12345") == "12345"
    assert n(None) == "" and n("nan") == ""


def test_parser_still_reads_old_spencers_layout(db):
    data = _old_layout(
        [["9000000001", "Old Rider", "South City", 12, "1,200"]],
        [["O1", "9000000001", 300]],
    )
    res = parse_file("Spencer's", data)
    assert res.matched_columns["rider_id"] == "Rider id"
    assert res.matched_columns["payout"] == "Total Payable Amount"
    assert res.matched_columns["orders"] == "Delivered Orders"
    assert [r.rider_id for r in res.records] == ["9000000001"]
    assert res.records[0].payout == 1200 * 100
    assert res.records[0].hub == "South City"
    assert res.cod_lines[0].worker_code == "9000000001"
    assert res.cod_lines[0].hub is None


# ── 2. hub sync + output hub ─────────────────────────────────────────────────


def _set_hub(db, rider_id, hub):
    db.execute(
        "UPDATE rider_master SET hub=? WHERE rider_id=? AND company=?",
        (hub, rider_id, "Spencer's"),
    )


def test_payout_file_updates_roster_hub_and_output_uses_it(db):
    pid = make_person(db, "Pradip Ray")
    make_rider(db, pid, "6295515978", "Spencer's", "Pradip Ray")
    _set_hub(db, "6295515978", "Old Hub")
    pid2 = make_person(db, "No Hub Col")
    make_rider(db, pid2, "8355007139", "Spencer's", "Rajbir")
    _set_hub(db, "8355007139", "Keeps")
    db.commit()
    data = _new_layout(
        [
            [6295515978, "Pradip Ray", "FULL_TIME", "Tamal", "h069", "NTS", 300, 1, 1, 1000],
            [8355007139, "Rajbir", "FULL_TIME", "Gorakhpur", "s055", None, 282, 1, 1, 2000],
        ],
        [],
    )
    # Dry run: the file's hub shows on the row but the roster is untouched.
    r = process_cycle("Spencer's", date(2026, 8, 10), date(2026, 8, 16), data, commit=False)
    rows = {x.rider_id: x for x in r.pay_rows + r.dues_rows}
    assert rows["6295515978"].hub == "NTS"
    assert rows["8355007139"].hub == "Keeps"  # blank store in the file → roster value
    assert [(u["rider_id"], u["old_hub"], u["new_hub"]) for u in r.hub_updates] == [
        ("6295515978", "Old Hub", "NTS")
    ]
    assert (
        db.execute(
            "SELECT hub FROM rider_master WHERE rider_id='6295515978' AND company=?",
            ("Spencer's",),
        ).fetchone()[0]
        == "Old Hub"
    )

    # Commit: the roster follows the file.
    r = process_cycle("Spencer's", date(2026, 8, 10), date(2026, 8, 16), data, commit=True)
    assert r.committed
    hubs = {
        row[0]: row[1]
        for row in db.execute(
            "SELECT rider_id, hub FROM rider_master WHERE company=?", ("Spencer's",)
        ).fetchall()
    }
    assert hubs == {"6295515978": "NTS", "8355007139": "Keeps"}


# ── 3. HOLD sheet ────────────────────────────────────────────────────────────


def _sheet_rows(xlsx_bytes: bytes, name: str) -> list[list]:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    out = []
    for r in wb[name].iter_rows(values_only=True):
        cells = list(r)
        while cells and cells[-1] is None:
            cells.pop()
        out.append(cells)
    return out


def test_hold_sheet_separates_cod_riders_not_in_payout(db):
    pid = make_person(db, "Pradip Ray")
    make_rider(db, pid, "6295515978", "Spencer's", "Pradip Ray")
    pid2 = make_person(db, "Anish Biswas")
    make_rider(db, pid2, "7439823275", "Spencer's", "Anish Biswas")
    _set_hub(db, "7439823275", "Roster Hub")
    db.commit()
    data = _new_layout(
        [[6295515978, "Pradip Ray", "FULL_TIME", "Tamal", "h069", "NTS", 300, 1, 1, 1000]],
        [
            # in the payout → held
            [3008272746, "H069", 6295515978, "Pradip Ray", 349, "Pending", "COD"],
            # on the roster but NOT in the payout
            [3008273151, "H049", 7439823275, "Anish Biswas", 609, "Pending", "COD"],
            [3008273152, "H049", 7439823275, "Anish Biswas", 583, "Pending", "COD"],
            # not on the roster at all, only the COD sheet knows them
            [3008270000, "H012", 9999999999, "Stranger", 250, "Pending", "COD"],
            # settled → not a hold
            [3008270001, "H012", 9999999999, "Stranger", 999, "Settled", "COD"],
        ],
    )
    r = process_cycle("Spencer's", date(2026, 8, 10), date(2026, 8, 16), data, commit=True)
    assert r.file_rider_ids == ["6295515978"]
    hold = {h["rider_id"]: h for h in r.hold_rows}
    assert hold["6295515978"]["in_payout"] is True
    assert hold["6295515978"]["hub"] == "NTS"  # payout sheet beats the COD hub code
    assert hold["7439823275"] == {
        "rider_id": "7439823275",
        "amount": (609 + 583) * 100,
        "name": "Anish Biswas",
        "hub": "H049",  # the file's hub, not the roster's
        "in_payout": False,
    }
    assert hold["9999999999"]["name"] == "Stranger" and hold["9999999999"]["hub"] == "H012"
    assert hold["9999999999"]["amount"] == 250 * 100

    rows = _sheet_rows(build_output(r).getvalue(), "HOLD")
    assert rows[0] == ["Rider ID", "Name", "Hub", "COD Total", "In Payout"]
    # In the payout → hub as the payout sheet states it (roster just synced).
    assert rows[1] == ["6295515978", "Pradip Ray", "NTS", 349.0, "Yes"]
    assert rows[2][1] == "TOTAL"
    # Second block: NOT in payout, own title + header, hubs from the COD sheet.
    assert rows[4][0] == NOT_IN_PAYOUT_TITLE
    assert rows[5] == ["Rider ID", "Name", "Hub", "COD Total", "In Payout"]
    block = rows[6:8]
    assert sorted(block, key=lambda x: x[0]) == [
        ["7439823275", "Anish Biswas", "H049", 1192.0, "No"],
        ["9999999999", "Stranger", "H012", 250.0, "No"],
    ]
    assert rows[8][1] == "TOTAL"
    # Line items carry name + hub too.
    li = rows.index(["Line Items"])
    assert rows[li + 1][:3] == ["Rider ID", "Name", "Hub"]
    lines = {(x[0], x[3]): x for x in rows[li + 2 :]}
    assert lines[("9999999999", "3008270000")][1:3] == ["Stranger", "H012"]
    assert lines[("7439823275", "3008273151")][2] == "H049"
