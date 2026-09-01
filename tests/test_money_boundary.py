"""Rupee/paise boundary regressions (2026-09-01 review).

The contract is "integer paise inside, rupees at the edge". Five places broke it:
COD clearance posted rupees into the paise ledger, the bank-MIS parser returned
rupee floats that were stored in a paise column, the dashboard Excel export's
add_styled_sheet ignored money_cols, and MONEY_KEYS lacked keys the UI renders.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from payout.exports import add_styled_sheet
from payout.money import MONEY_KEYS, rupeeize
from payout.parsers.bank_mis import _column_indexes, _to_amount
from tests.conftest import make_person


def test_rupeeize_covers_the_dashboard_and_ev_rent_keys():
    for key in ("owed", "held", "collected_current", "rolled_recovered_later", "rolled_forward_net"):
        assert key in MONEY_KEYS, key
    out = rupeeize({"owed": 125000, "held": 50000, "days": 7, "rows": [{"collected_current": 250000}]})
    assert out == {"owed": 1250.0, "held": 500.0, "days": 7, "rows": [{"collected_current": 2500.0}]}


def test_add_styled_sheet_converts_money_columns():
    wb = Workbook()
    add_styled_sheet(
        wb, sheet_name="Money Flow", headers=["Company", "Gross", "Riders"],
        rows=[["Blitz", 2290000, 12]], numeric_cols=[2, 3], money_cols=[2],
    )
    ws = wb["Money Flow"]
    assert ws.cell(row=2, column=2).value == 22900.0   # was 2290000 (paise)
    assert ws.cell(row=2, column=3).value == 12         # counts untouched


def test_bank_mis_amount_is_paise():
    assert _to_amount("1,250.50") == 125050
    assert _to_amount("800") == 80000
    assert _to_amount("") == 0


def test_bank_mis_wrapped_headers_do_not_collide():
    hdr = ["Sr", "Pymt_Mode", "Beneficia\nry Name", "Beneficia\nry Account No",
           "Bene_IFS\nC_Code", "Amount", "Remark", "Pymt_Da\nte", "Status",
           "Custome\nr Ref No", "UTR No"]
    cols = _column_indexes(hdr)
    assert cols["bene_name"] == 2
    assert cols["bene_acc"] == 3        # used to be 2: account read as the name
    assert cols["ifsc"] == 4 and cols["amount"] == 5 and cols["ref"] == 9 and cols["utr"] == 10


@pytest.fixture
def admin_client(db):
    pytest.importorskip("fastapi")
    from fastapi.testclient import TestClient

    from payout.api import ratelimit
    from payout.api.app import app
    from payout.auth import hash_password

    db.execute(
        "INSERT INTO users (email, password_hash, role, is_active) VALUES (?,?,?,1)",
        ("adm@t.test", hash_password("Admin-pass-1"), "admin"),
    )
    db.commit()
    ratelimit.reset()
    with TestClient(app) as c:
        r = c.post("/api/auth/login", data={"username": "adm@t.test", "password": "Admin-pass-1"})
        assert r.status_code == 200
        yield c


def test_cod_clearance_posts_paise_to_the_ledger(db, admin_client):
    pid = make_person(db, "C", balance=0)
    db.execute(
        "INSERT INTO cod_holds (cycle_start, cycle_end, company, rider_id, person_id, "
        "worker_code, amount, source) VALUES ('2026-06-01','2026-06-07','Myntra','C1',?,'C1',50000,'myntra_column')",
        (pid,),
    )
    db.commit()
    r = admin_client.post("/api/cod/clear", json={"person_id": pid, "ledger_amount": 500})
    assert r.status_code == 200, r.text
    stored = db.execute(
        "SELECT amount FROM transactions WHERE person_id=? AND event_type='ADJUSTMENT'", (pid,)
    ).fetchone()[0]
    assert stored == 50000                       # ₹500 -> 50,000 paise (was 500)
    assert db.execute("SELECT current_balance FROM balances WHERE person_id=?", (pid,)).fetchone()[0] == 50000
    # and the JSON edge converts back to rupees
    assert r.json()["new_balance"] == 500.0
